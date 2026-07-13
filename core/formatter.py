
"""
formatter.py
------------------------------------
Format Excel Report
"""

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter


TITLE_FILL = PatternFill(
    fill_type="solid",
    start_color="1F4E78"
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="D9EAD3"
)

TITLE_FONT = Font(
    bold=True,
    color="FFFFFF",
    size=16
)

HEADER_FONT = Font(
    bold=True
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def format_report(workbook):
    """
    Apply formatting to Management Report.
    """

    sheet = workbook["Management Report"]

    # -------------------------
    # Report Title
    # -------------------------

    sheet.merge_cells("A1:B1")

    title = sheet["A1"]

    title.font = TITLE_FONT
    title.fill = TITLE_FILL
    title.alignment = Alignment(horizontal="center")

    # -------------------------
    # Header Rows
    # -------------------------

    header_rows = [3, 11, 25, 38]

    for row in header_rows:

        cell = sheet[f"A{row}"]

        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # -------------------------
    # Borders
    # -------------------------

    for row in sheet.iter_rows():

        for cell in row:

            if cell.value is not None:

                cell.border = THIN_BORDER

    # -------------------------
    # Currency Formatting
    # -------------------------

    for row in range(4, sheet.max_row + 1):

        value = sheet[f"B{row}"].value

        if isinstance(value, (int, float)):

            sheet[f"B{row}"].number_format = '#,##0.00'

    # -------------------------
    # Auto Width
    # -------------------------

    for col_num in range(1, sheet.max_column + 1):

        column_letter = get_column_letter(col_num)

        max_length = 0

        for row in range(1, sheet.max_row + 1):

            cell = sheet.cell(row=row, column=col_num)

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        sheet.column_dimensions[column_letter].width = max_length + 3

    # -------------------------
    # Freeze Pane
    # -------------------------

    sheet.freeze_panes = "A4"

